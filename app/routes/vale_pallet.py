"""
Rotas para gerenciamento de Vale Pallet
"""
from flask import Blueprint, render_template, redirect, url_for, flash, request
from flask_login import login_required, current_user
from app import db
from app.models import ValePallet, Empresa, TipoEmpresa, Motorista
from app.forms import ValePalletForm
from app.utils.whatsapp import enviar_whatsapp_vale_criado

bp = Blueprint('vale_pallet', __name__, url_prefix='/vale-pallet')


@bp.route('/')
@login_required
def listar():
    """Lista todos os vales pallet que o usuário pode ver (com filtros)"""
    from datetime import datetime
    
    page = request.args.get('page', 1, type=int)
    per_page = 15  # 15 vales por página
    
    # Filtros
    status_filtro = request.args.get('status', '')
    destinatario_filtro = request.args.get('destinatario', '', type=int)
    data_vencimento_inicio = request.args.get('data_vencimento_inicio', '')
    data_vencimento_fim = request.args.get('data_vencimento_fim', '')
    
    # Vales das empresas que o usuário pode ver
    empresas_visiveis = current_user.empresas_visiveis()
    empresas_ids = [e.id for e in empresas_visiveis]
    
    # Query base
    query = ValePallet.query.filter(
        (ValePallet.cliente_id.in_(empresas_ids)) |
        (ValePallet.transportadora_id.in_(empresas_ids)) |
        (ValePallet.destinatario_id.in_(empresas_ids))
    )
    
    # Aplicar filtro de status
    if status_filtro:
        query = query.filter(ValePallet.status == status_filtro)
    
    # Aplicar filtro de destinatário
    if destinatario_filtro:
        query = query.filter(ValePallet.destinatario_id == destinatario_filtro)
    
    # Aplicar filtro de data de vencimento (início)
    if data_vencimento_inicio:
        try:
            data_inicio = datetime.strptime(data_vencimento_inicio, '%Y-%m-%d').date()
            query = query.filter(ValePallet.data_vencimento >= data_inicio)
        except ValueError:
            pass
    
    # Aplicar filtro de data de vencimento (fim)
    if data_vencimento_fim:
        try:
            data_fim = datetime.strptime(data_vencimento_fim, '%Y-%m-%d').date()
            query = query.filter(ValePallet.data_vencimento <= data_fim)
        except ValueError:
            pass
    
    # Ordenar e paginar
    vales = query.order_by(ValePallet.data_criacao.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    # Buscar destinatários para o filtro
    tipo_destinatario = TipoEmpresa.query.filter_by(nome='Destinatário').first()
    destinatarios = []
    if tipo_destinatario:
        destinatarios = [e for e in empresas_visiveis if e.tipo_empresa_id == tipo_destinatario.id]
    else:
        destinatarios = empresas_visiveis
    
    # Status disponíveis
    status_opcoes = [
        ('pendente_entrega', 'Pendente de Entrega'),
        ('entrega_realizada', 'Entrega Realizada'),
        ('entrega_concluida', 'Entrega Concluída'),
        ('cancelado', 'Cancelado')
    ]
    
    return render_template('vale_pallet/listar.html',
                          vales=vales,
                          destinatarios=destinatarios,
                          status_opcoes=status_opcoes,
                          status_filtro=status_filtro,
                          destinatario_filtro=destinatario_filtro,
                          data_vencimento_inicio=data_vencimento_inicio,
                          data_vencimento_fim=data_vencimento_fim)


@bp.route('/novo', methods=['GET', 'POST'])
@login_required
def novo():
    """Cadastra um novo vale pallet"""
    form = ValePalletForm()
    
    # Carregar empresas que o usuário pode ver, separadas por tipo
    empresas_visiveis = current_user.empresas_visiveis()
    
    tipo_cliente = TipoEmpresa.query.filter_by(nome='Cliente').first()
    tipo_transportadora = TipoEmpresa.query.filter_by(nome='Transportadora').first()
    tipo_destinatario = TipoEmpresa.query.filter_by(nome='Destinatário').first()
    
    # Filtrar empresas por tipo
    clientes = [e for e in empresas_visiveis if tipo_cliente and e.tipo_empresa_id == tipo_cliente.id]
    transportadoras = [e for e in empresas_visiveis if tipo_transportadora and e.tipo_empresa_id == tipo_transportadora.id]
    destinatarios = [e for e in empresas_visiveis if tipo_destinatario and e.tipo_empresa_id == tipo_destinatario.id]
    
    # Se não houver tipos cadastrados, mostrar todas as empresas
    if not tipo_cliente:
        clientes = empresas_visiveis
    if not tipo_transportadora:
        transportadoras = empresas_visiveis
    if not tipo_destinatario:
        destinatarios = empresas_visiveis
    
    form.cliente_id.choices = [(0, 'Selecione um cliente')] + [
        (e.id, e.razao_social) for e in clientes
    ]
    form.transportadora_id.choices = [(0, 'Selecione uma transportadora')] + [
        (e.id, e.razao_social) for e in transportadoras
    ]
    form.destinatario_id.choices = [(0, 'Selecione um destinatário')] + [
        (e.id, e.razao_social) for e in destinatarios
    ]
    
    # Carregar motoristas ativos
    motoristas = Motorista.query.filter_by(ativo=True).all()
    form.motorista_id.choices = [(0, 'Selecione um motorista (opcional)')] + [
        (m.id, f"{m.nome} - {m.placa_caminhao}") for m in motoristas
    ]
    
    if form.validate_on_submit():
        # Verificar se as empresas existem e o usuário pode vê-las
        cliente = Empresa.query.get(form.cliente_id.data)
        transportadora = Empresa.query.get(form.transportadora_id.data)
        destinatario = Empresa.query.get(form.destinatario_id.data)
        
        if not cliente or not transportadora or not destinatario:
            flash('Uma ou mais empresas selecionadas não foram encontradas!', 'danger')
            return render_template('vale_pallet/form.html', form=form, titulo='Novo Vale Pallet')
        
        if not (current_user.pode_ver_empresa(cliente.id) and 
                current_user.pode_ver_empresa(transportadora.id) and 
                current_user.pode_ver_empresa(destinatario.id)):
            flash('Você não tem permissão para criar vale pallet com uma ou mais empresas selecionadas!', 'danger')
            return redirect(url_for('vale_pallet.listar'))
        
        # Gerar PIN único
        pin = ValePallet.gerar_pin()
        
        # Criar novo vale pallet
        vale = ValePallet(
            cliente_id=form.cliente_id.data,
            transportadora_id=form.transportadora_id.data,
            destinatario_id=form.destinatario_id.data,
            motorista_id=form.motorista_id.data if form.motorista_id.data != 0 else None,
            quantidade_pallets=int(form.quantidade_pallets.data),
            numero_documento=form.numero_documento.data,
            data_vencimento=form.data_vencimento.data,
            pin=pin,
            status='pendente_entrega',
            criado_por_id=current_user.id
        )
        
        db.session.add(vale)
        db.session.commit()
        
        # Enviar email para o destinatário
        from app.utils.email_service import enviar_email_vale_pallet
        resultado_email = enviar_email_vale_pallet(vale.id, current_user.id)
        
        # Enviar WhatsApp para o motorista (se selecionado)
        mensagem_sucesso = 'Vale Pallet criado com sucesso!'
        
        if resultado_email['sucesso']:
            mensagem_sucesso += f'. {resultado_email["emails_enviados"]} email(s) enviado(s)'
        else:
            mensagem_sucesso += f'. Aviso: {resultado_email["mensagem"]}'
        
        if vale.motorista_id:
            motorista = Motorista.query.get(vale.motorista_id)
            if motorista and motorista.celular:
                enviado = enviar_whatsapp_vale_criado(motorista, vale)
                if enviado:
                    mensagem_sucesso += '. WhatsApp enviado para o motorista'
                else:
                    mensagem_sucesso += '. Erro ao enviar WhatsApp para o motorista'
            else:
                mensagem_sucesso += '. Motorista sem celular cadastrado'
        
        flash(mensagem_sucesso, 'success' if resultado_email['sucesso'] else 'warning')
        
        return redirect(url_for('vale_pallet.visualizar', id=vale.id))
    
    return render_template('vale_pallet/form.html', form=form, titulo='Novo Vale Pallet')


@bp.route('/<int:id>')
@login_required
def visualizar(id):
    """Visualiza os detalhes de um vale pallet"""
    vale = ValePallet.query.get_or_404(id)
    
    # Verificar se o usuário pode ver este vale
    if not (current_user.pode_ver_empresa(vale.cliente_id) or 
            current_user.pode_ver_empresa(vale.transportadora_id) or 
            current_user.pode_ver_empresa(vale.destinatario_id)):
        flash('Você não tem permissão para visualizar este vale pallet!', 'danger')
        return redirect(url_for('vale_pallet.listar'))
    
    return render_template('vale_pallet/visualizar.html', vale=vale)


@bp.route('/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar(id):
    """Edita um vale pallet"""
    vale = ValePallet.query.get_or_404(id)
    
    # Verificar se o usuário pode editar este vale
    if not (current_user.pode_ver_empresa(vale.cliente_id) or 
            current_user.pode_ver_empresa(vale.transportadora_id) or 
            current_user.pode_ver_empresa(vale.destinatario_id)):
        flash('Você não tem permissão para editar este vale pallet!', 'danger')
        return redirect(url_for('vale_pallet.listar'))
    
    form = ValePalletForm(obj=vale)
    
    # Carregar empresas que o usuário pode ver, separadas por tipo
    empresas_visiveis = current_user.empresas_visiveis()
    
    tipo_cliente = TipoEmpresa.query.filter_by(nome='Cliente').first()
    tipo_transportadora = TipoEmpresa.query.filter_by(nome='Transportadora').first()
    tipo_destinatario = TipoEmpresa.query.filter_by(nome='Destinatário').first()
    
    # Filtrar empresas por tipo
    clientes = [e for e in empresas_visiveis if tipo_cliente and e.tipo_empresa_id == tipo_cliente.id]
    transportadoras = [e for e in empresas_visiveis if tipo_transportadora and e.tipo_empresa_id == tipo_transportadora.id]
    destinatarios = [e for e in empresas_visiveis if tipo_destinatario and e.tipo_empresa_id == tipo_destinatario.id]
    
    # Se não houver tipos cadastrados, mostrar todas as empresas
    if not tipo_cliente:
        clientes = empresas_visiveis
    if not tipo_transportadora:
        transportadoras = empresas_visiveis
    if not tipo_destinatario:
        destinatarios = empresas_visiveis
    
    form.cliente_id.choices = [(0, 'Selecione um cliente')] + [
        (e.id, e.razao_social) for e in clientes
    ]
    form.transportadora_id.choices = [(0, 'Selecione uma transportadora')] + [
        (e.id, e.razao_social) for e in transportadoras
    ]
    form.destinatario_id.choices = [(0, 'Selecione um destinatário')] + [
        (e.id, e.razao_social) for e in destinatarios
    ]
    
    if form.validate_on_submit():
        # Atualizar vale pallet (mantém o PIN original)
        vale.cliente_id = form.cliente_id.data
        vale.transportadora_id = form.transportadora_id.data
        vale.destinatario_id = form.destinatario_id.data
        vale.quantidade_pallets = int(form.quantidade_pallets.data)
        vale.numero_documento = form.numero_documento.data
        vale.data_vencimento = form.data_vencimento.data
        
        db.session.commit()
        
        flash(f'Vale Pallet atualizado com sucesso!', 'success')
        return redirect(url_for('vale_pallet.visualizar', id=vale.id))
    
    # Preencher os valores atuais
    if request.method == 'GET':
        form.cliente_id.data = vale.cliente_id
        form.transportadora_id.data = vale.transportadora_id
        form.destinatario_id.data = vale.destinatario_id
        form.quantidade_pallets.data = str(vale.quantidade_pallets)
        form.data_vencimento.data = vale.data_vencimento
    
    return render_template('vale_pallet/form.html', form=form, vale=vale, titulo='Editar Vale Pallet')


@bp.route('/<int:id>/cancelar', methods=['POST'])
@login_required
def cancelar(id):
    """Cancela um vale pallet"""
    vale = ValePallet.query.get_or_404(id)
    
    # Verificar se o usuário pode cancelar este vale
    if not (current_user.pode_ver_empresa(vale.cliente_id) or 
            current_user.pode_ver_empresa(vale.transportadora_id) or 
            current_user.pode_ver_empresa(vale.destinatario_id)):
        flash('Você não tem permissão para cancelar este vale pallet!', 'danger')
        return redirect(url_for('vale_pallet.listar'))
    
    vale.status = 'cancelado'
    db.session.commit()
    
    flash(f'Vale Pallet cancelado com sucesso!', 'success')
    return redirect(url_for('vale_pallet.visualizar', id=vale.id))


@bp.route('/<int:id>/finalizar', methods=['POST'])
@login_required
def finalizar(id):
    """Finaliza um vale pallet"""
    vale = ValePallet.query.get_or_404(id)
    
    # Verificar se o usuário pode finalizar este vale
    if not (current_user.pode_ver_empresa(vale.cliente_id) or 
            current_user.pode_ver_empresa(vale.transportadora_id) or 
            current_user.pode_ver_empresa(vale.destinatario_id)):
        flash('Você não tem permissão para finalizar este vale pallet!', 'danger')
        return redirect(url_for('vale_pallet.listar'))
    
    vale.status = 'finalizado'
    db.session.commit()
    
    flash(f'Vale Pallet finalizado com sucesso!', 'success')
    return redirect(url_for('vale_pallet.visualizar', id=vale.id))


@bp.route('/buscar-pin', methods=['GET', 'POST'])
@login_required
def buscar_pin():
    """Busca um vale pallet pelo PIN"""
    vale = None
    pin = request.args.get('pin') or request.form.get('pin')
    
    if pin:
        vale = ValePallet.query.filter_by(pin=pin).first()
        
        if vale:
            # Verificar se o usuário pode ver este vale
            if not (current_user.pode_ver_empresa(vale.cliente_id) or 
                    current_user.pode_ver_empresa(vale.transportadora_id) or 
                    current_user.pode_ver_empresa(vale.destinatario_id)):
                flash('Você não tem permissão para visualizar este vale pallet!', 'danger')
                return redirect(url_for('vale_pallet.buscar_pin'))
            
            return redirect(url_for('vale_pallet.visualizar', id=vale.id))
        else:
            flash(f'Vale Pallet com PIN "{pin}" não encontrado!', 'warning')
    
    return render_template('vale_pallet/buscar_pin.html')


@bp.route('/exportar-excel')
@login_required
def exportar_excel():
    """Exporta lista de vales para Excel (respeitando filtros)"""
    from datetime import datetime
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Filtros (mesmos da listagem)
    status_filtro = request.args.get('status', '')
    destinatario_filtro = request.args.get('destinatario', '', type=int)
    data_vencimento_inicio = request.args.get('data_vencimento_inicio', '')
    data_vencimento_fim = request.args.get('data_vencimento_fim', '')
    
    # Vales das empresas que o usuário pode ver
    empresas_visiveis = current_user.empresas_visiveis()
    empresas_ids = [e.id for e in empresas_visiveis]
    
    # Query base (SEM paginação - todos os registros)
    query = ValePallet.query.filter(
        (ValePallet.cliente_id.in_(empresas_ids)) |
        (ValePallet.transportadora_id.in_(empresas_ids)) |
        (ValePallet.destinatario_id.in_(empresas_ids))
    )
    
    # Aplicar filtros
    if status_filtro:
        query = query.filter(ValePallet.status == status_filtro)
    
    if destinatario_filtro:
        query = query.filter(ValePallet.destinatario_id == destinatario_filtro)
    
    if data_vencimento_inicio:
        try:
            data_inicio = datetime.strptime(data_vencimento_inicio, '%Y-%m-%d').date()
            query = query.filter(ValePallet.data_vencimento >= data_inicio)
        except ValueError:
            pass
    
    if data_vencimento_fim:
        try:
            data_fim = datetime.strptime(data_vencimento_fim, '%Y-%m-%d').date()
            query = query.filter(ValePallet.data_vencimento <= data_fim)
        except ValueError:
            pass
    
    # Buscar TODOS os vales (sem paginação)
    vales = query.order_by(ValePallet.data_criacao.desc()).all()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vales Pallet"
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Cabeçalhos
    headers = [
        "PIN", "Documento", "Cliente", "Destinatário", "Transportadora",
        "Motorista", "Placa", "Quantidade Pallets", "Quantidade Devolvida",
        "Saldo", "Data Vencimento", "Data Criação", "Status"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Dados
    for row_num, vale in enumerate(vales, 2):
        ws.cell(row=row_num, column=1, value=vale.pin)
        ws.cell(row=row_num, column=2, value=vale.numero_documento)
        ws.cell(row=row_num, column=3, value=vale.cliente.razao_social if vale.cliente else "")
        ws.cell(row=row_num, column=4, value=vale.destinatario.razao_social if vale.destinatario else "")
        ws.cell(row=row_num, column=5, value=vale.transportadora.razao_social if vale.transportadora else "")
        ws.cell(row=row_num, column=6, value=vale.motorista.nome if vale.motorista else "")
        ws.cell(row=row_num, column=7, value=vale.motorista.placa_caminhao if vale.motorista else "")
        ws.cell(row=row_num, column=8, value=vale.quantidade_pallets)
        ws.cell(row=row_num, column=9, value=vale.quantidade_devolvida)
        ws.cell(row=row_num, column=10, value=vale.quantidade_pallets - vale.quantidade_devolvida)
        ws.cell(row=row_num, column=11, value=vale.data_vencimento.strftime('%d/%m/%Y') if vale.data_vencimento else "")
        ws.cell(row=row_num, column=12, value=vale.data_criacao.strftime('%d/%m/%Y %H:%M') if vale.data_criacao else "")
        ws.cell(row=row_num, column=13, value=vale.get_status_display())
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Salvar em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome do arquivo com data/hora
    nome_arquivo = f"vales_pallet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_arquivo
    )



@bp.route('/<int:id>/imprimir-pdf')
@login_required
def imprimir_pdf(id):
    """Gera PDF do vale pallet com QR Code de validação - Meia folha A4"""
    from flask import make_response
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from io import BytesIO
    from app.utils.qrcode_utils import gerar_qrcode_vale
    import base64
    
    vale = ValePallet.query.get_or_404(id)
    
    # Verificar se o usuário pode ver este vale
    if not (current_user.pode_ver_empresa(vale.cliente_id) or 
            current_user.pode_ver_empresa(vale.transportadora_id) or 
            current_user.pode_ver_empresa(vale.destinatario_id)):
        flash('Você não tem permissão para visualizar este vale pallet!', 'danger')
        return redirect(url_for('vale_pallet.listar'))
    
    # Criar buffer para PDF
    buffer = BytesIO()
    
    # Criar PDF - A4 completo (mas vamos usar só metade)
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Definir altura de meia folha (A4 dividido ao meio horizontalmente)
    half_height = height / 2
    
    # Começar do topo da meia folha superior
    y_start = height - 1*cm
    y = y_start
    
    # ==================== CABEÇALHO ====================
    # Logo do lado esquerdo
    from reportlab.lib.utils import ImageReader
    import os
    logo_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'static', 'logo.png')
    
    try:
        logo = ImageReader(logo_path)
        # Logo com 2cm de altura, proporcional
        logo_height = 1.5*cm
        logo_width = logo_height * 2  # Ajustar proporção conforme necessário
        p.drawImage(logo, 1*cm, y - logo_height - 0.3*cm, width=logo_width, height=logo_height, preserveAspectRatio=True, mask='auto')
    except:
        # Se logo não existir, mostrar texto
        p.setFont("Helvetica-Bold", 14)
        p.setFillColor(colors.HexColor('#2dce89'))
        p.drawString(1*cm, y - 1*cm, "ePALLET")
    
    # Texto do lado direito
    p.setFont("Helvetica", 9)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawRightString(width - 1*cm, y - 0.8*cm, "Documento com assinatura digital")
    
    # Linha separadora
    p.setStrokeColor(colors.HexColor('#2dce89'))
    p.setLineWidth(1.5)
    p.line(1*cm, y - 2*cm, width - 1*cm, y - 2*cm)
    
    y -= 2.5*cm
    
    # ==================== INFORMAÇÕES DO VALE ====================
    p.setFillColor(colors.black)
    
    # Seção: Informações do Vale Pallet
    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.HexColor('#5e72e4'))
    p.drawString(1*cm, y, "📄 Informações do Vale Pallet")
    p.setFillColor(colors.black)
    y -= 0.5*cm
    
    # Número do Documento e Status na mesma linha
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawString(1*cm, y, "Documento:")
    p.setFont("Helvetica-Bold", 8)
    p.setFillColor(colors.black)
    p.drawString(3*cm, y, vale.numero_documento)
    
    # Status do lado direito
    status_display = {
        'pendente_entrega': 'Pendente de Entrega',
        'entrega_realizada': 'Entrega Realizada',
        'entrega_concluida': 'Entrega Concluída',
        'cancelado': 'Cancelado'
    }.get(vale.status, vale.status)
    
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawString(width/2, y, "Status:")
    p.setFont("Helvetica-Bold", 7)
    p.setFillColor(colors.HexColor('#5e72e4'))
    p.drawString(width/2 + 1.5*cm, y, status_display)
    p.setFillColor(colors.black)
    y -= 0.7*cm
    
    # Seção: Empresas Envolvidas
    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.HexColor('#11cdef'))
    p.drawString(1*cm, y, "🏢 Empresas Envolvidas")
    p.setFillColor(colors.black)
    y -= 0.5*cm
    
    # Cliente
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawString(1*cm, y, "Cliente:")
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.black)
    cliente_nome = vale.cliente.razao_social[:60] if vale.cliente else '-'
    p.drawString(3*cm, y, cliente_nome)
    y -= 0.4*cm
    
    # Destinatário
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawString(1*cm, y, "Destinatário:")
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.black)
    dest_nome = vale.destinatario.razao_social[:60] if vale.destinatario else '-'
    p.drawString(3*cm, y, dest_nome)
    y -= 0.4*cm
    
    # Transportadora
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawString(1*cm, y, "Transportadora:")
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.black)
    transp_nome = vale.transportadora.razao_social[:60] if vale.transportadora else '-'
    p.drawString(3*cm, y, transp_nome)
    y -= 0.7*cm
    
    # Seção: Informações de Entrega
    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.HexColor('#fb6340'))
    p.drawString(1*cm, y, "🚚 Informações de Entrega")
    p.setFillColor(colors.black)
    y -= 0.5*cm
    
    # Motorista, Placa e Celular na mesma linha
    if vale.motorista:
        motorista_nome = vale.motorista.nome[:25] if vale.motorista.nome else 'Não informado'
        placa = vale.motorista.placa_caminhao if vale.motorista.placa_caminhao else '-'
        celular = vale.motorista.celular if vale.motorista.celular else '-'
        
        p.setFont("Helvetica", 7)
        p.setFillColor(colors.HexColor('#6c757d'))
        p.drawString(1*cm, y, "Motorista:")
        p.setFont("Helvetica", 7)
        p.setFillColor(colors.black)
        p.drawString(3*cm, y, motorista_nome)
        
        p.setFont("Helvetica", 7)
        p.setFillColor(colors.HexColor('#6c757d'))
        p.drawString(8*cm, y, "Placa:")
        p.setFont("Helvetica", 7)
        p.setFillColor(colors.black)
        p.drawString(9*cm, y, placa)
        
        p.setFont("Helvetica", 7)
        p.setFillColor(colors.HexColor('#6c757d'))
        p.drawString(11*cm, y, "Celular:")
        p.setFont("Helvetica", 7)
        p.setFillColor(colors.black)
        p.drawString(12.5*cm, y, celular)
    else:
        p.setFont("Helvetica", 7)
        p.setFillColor(colors.HexColor('#6c757d'))
        p.drawString(1*cm, y, "Motorista:")
        p.setFont("Helvetica", 7)
        p.setFillColor(colors.black)
        p.drawString(3*cm, y, "Não informado")
    
    y -= 0.7*cm
    
    # Seção: Quantidade de Pallets
    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.HexColor('#2dce89'))
    p.drawString(1*cm, y, "📦 Quantidade de Pallets")
    p.setFillColor(colors.black)
    y -= 0.5*cm
    
    # Quantidade Total
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawString(1*cm, y, "Quantidade Total:")
    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.HexColor('#5e72e4'))
    p.drawString(4*cm, y, f"{vale.quantidade_pallets} pallets")
    p.setFillColor(colors.black)
    y -= 0.7*cm
    
    # Seção: Datas
    p.setFont("Helvetica-Bold", 9)
    p.setFillColor(colors.HexColor('#f5365c'))
    p.drawString(1*cm, y, "📅 Datas")
    p.setFillColor(colors.black)
    y -= 0.5*cm
    
    # Data de Criação e Vencimento na mesma linha
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawString(1*cm, y, "Criação:")
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.black)
    p.drawString(3*cm, y, vale.data_criacao.strftime('%d/%m/%Y %H:%M'))
    
    # Data de Vencimento
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.HexColor('#6c757d'))
    p.drawString(8*cm, y, "Vencimento:")
    p.setFont("Helvetica", 7)
    
    # Verificar se está vencido
    from datetime import datetime
    vencido = vale.data_vencimento < datetime.now().date()
    p.setFillColor(colors.HexColor('#f5365c') if vencido else colors.HexColor('#2dce89'))
    p.drawString(10.5*cm, y, vale.data_vencimento.strftime('%d/%m/%Y'))
    
    if vencido:
        p.setFont("Helvetica-Bold", 6)
        p.drawString(13*cm, y, "⚠ VENCIDO")
    
    # ==================== QR CODE ====================
    # Gerar QR Code
    qr_base64 = gerar_qrcode_vale(vale.id, vale.numero_documento, vale.pin, vale.data_criacao)
    
    from reportlab.lib.utils import ImageReader
    qr_image = ImageReader(BytesIO(base64.b64decode(qr_base64)))
    
    # Posicionar QR Code no canto inferior direito
    qr_size = 3.5*cm
    qr_x = width - 1.5*cm - qr_size
    qr_y = half_height + 0.8*cm
    
    # Borda ao redor do QR Code
    p.setStrokeColor(colors.HexColor('#2dce89'))
    p.setLineWidth(2)
    p.rect(qr_x - 0.2*cm, qr_y - 0.2*cm, qr_size + 0.4*cm, qr_size + 0.4*cm, fill=False, stroke=True)
    
    p.drawImage(qr_image, qr_x, qr_y, width=qr_size, height=qr_size)
    
    # Texto abaixo do QR Code
    p.setFont("Helvetica-Bold", 7)
    p.setFillColor(colors.HexColor('#2dce89'))
    p.drawCentredString(qr_x + qr_size/2, qr_y - 0.4*cm, "Escaneie para validar")
    
    # ==================== RODAPÉ ====================
    p.setStrokeColor(colors.HexColor('#2dce89'))
    p.setLineWidth(1)
    p.line(1.5*cm, half_height + 0.5*cm, width - 1.5*cm, half_height + 0.5*cm)
    
    p.setFont("Helvetica", 6)
    p.setFillColor(colors.grey)
    p.drawCentredString(width/2, half_height + 0.2*cm, f"Sistema ePallet - Gestão Inteligente de Pallets | Gerado por: {current_user.username}")
    
    # ==================== LINHA DE CORTE ====================
    # Linha tracejada no meio da folha
    p.setStrokeColor(colors.grey)
    p.setLineWidth(0.5)
    p.setDash(3, 3)
    p.line(0, half_height, width, half_height)
    
    # Texto "CORTE AQUI"
    p.setFont("Helvetica", 7)
    p.setFillColor(colors.grey)
    p.drawCentredString(width/2, half_height - 0.3*cm, "✂ CORTE AQUI ✂")
    
    # Finalizar PDF
    p.showPage()
    p.save()
    
    # Retornar PDF
    buffer.seek(0)
    response = make_response(buffer.getvalue())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = f'inline; filename=vale_pallet_{vale.numero_documento}.pdf'
    
    return response
