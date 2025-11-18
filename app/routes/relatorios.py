from flask import Blueprint, render_template, request
from flask_login import login_required, current_user
from app.models import ValePallet, DevolucaoPallet, Empresa, TipoEmpresa, Motorista
from sqlalchemy import or_, and_
from app.utils.auditoria import log_acesso_tela

bp = Blueprint('relatorios', __name__, url_prefix='/relatorios')


@bp.route('/movimentacao')
@login_required
def movimentacao():
    """Relatório de movimentação de vales pallet e devoluções com filtros"""
    log_acesso_tela('relatorios', 'Relatório de Movimentação')
    
    # Obter filtros da query string
    filtro_tipo_relatorio = request.args.get('tipo_relatorio', 'vales')  # vales, devolucoes, consolidado
    filtro_tipo = request.args.get('tipo_empresa', '')
    filtro_empresa = request.args.get('empresa', '')
    filtro_motorista = request.args.get('motorista', '')
    filtro_documento = request.args.get('documento', '')
    filtro_status = request.args.get('status', '')
    
    vales = []
    devolucoes = []
    
    # Buscar Vales Pallet
    if filtro_tipo_relatorio in ['vales', 'consolidado']:
        query_vales = ValePallet.query
        
        # Aplicar filtros em vales
        if filtro_tipo:
            if filtro_tipo == 'cliente':
                query_vales = query_vales.join(ValePallet.cliente).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome == 'Cliente'
                )
            elif filtro_tipo == 'transportadora':
                query_vales = query_vales.join(ValePallet.transportadora).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome == 'Transportadora'
                )
            elif filtro_tipo == 'destinatario':
                query_vales = query_vales.join(ValePallet.destinatario).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome.in_(['Destinatário', 'Destinatario'])
                )
        
        if filtro_empresa:
            # Filtrar por nome de empresa
            empresas_filtradas = Empresa.query.filter(
                or_(
                    Empresa.razao_social.ilike(f'%{filtro_empresa}%'),
                    Empresa.nome_fantasia.ilike(f'%{filtro_empresa}%')
                )
            ).all()
            empresas_ids = [e.id for e in empresas_filtradas]
            
            if empresas_ids:
                query_vales = query_vales.filter(
                    or_(
                        ValePallet.cliente_id.in_(empresas_ids),
                        ValePallet.transportadora_id.in_(empresas_ids),
                        ValePallet.destinatario_id.in_(empresas_ids)
                    )
                )
        
        if filtro_motorista:
            query_vales = query_vales.join(ValePallet.motorista).filter(
                Motorista.nome.ilike(f'%{filtro_motorista}%')
            )
        
        if filtro_documento:
            query_vales = query_vales.filter(
                ValePallet.numero_documento.ilike(f'%{filtro_documento}%')
            )
        
        if filtro_status:
            query_vales = query_vales.filter(ValePallet.status == filtro_status)
        
        vales = query_vales.order_by(ValePallet.data_criacao.desc()).all()
    
    # Buscar Devoluções
    if filtro_tipo_relatorio in ['devolucoes', 'consolidado']:
        query_devolucoes = DevolucaoPallet.query
        
        # Aplicar filtros em devoluções
        if filtro_tipo:
            if filtro_tipo == 'cliente':
                query_devolucoes = query_devolucoes.join(DevolucaoPallet.cliente).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome == 'Cliente'
                )
            elif filtro_tipo == 'transportadora':
                query_devolucoes = query_devolucoes.join(DevolucaoPallet.transportadora).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome == 'Transportadora'
                )
            elif filtro_tipo == 'destinatario':
                query_devolucoes = query_devolucoes.join(DevolucaoPallet.destinatario).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome.in_(['Destinatário', 'Destinatario'])
                )
        
        if filtro_empresa:
            # Filtrar por nome de empresa
            empresas_filtradas = Empresa.query.filter(
                or_(
                    Empresa.razao_social.ilike(f'%{filtro_empresa}%'),
                    Empresa.nome_fantasia.ilike(f'%{filtro_empresa}%')
                )
            ).all()
            empresas_ids = [e.id for e in empresas_filtradas]
            
            if empresas_ids:
                query_devolucoes = query_devolucoes.filter(
                    or_(
                        DevolucaoPallet.cliente_id.in_(empresas_ids),
                        DevolucaoPallet.transportadora_id.in_(empresas_ids),
                        DevolucaoPallet.destinatario_id.in_(empresas_ids)
                    )
                )
        
        if filtro_motorista:
            query_devolucoes = query_devolucoes.join(DevolucaoPallet.motorista).filter(
                Motorista.nome.ilike(f'%{filtro_motorista}%')
            )
        
        if filtro_status:
            query_devolucoes = query_devolucoes.filter(DevolucaoPallet.status == filtro_status)
        
        devolucoes = query_devolucoes.order_by(DevolucaoPallet.criado_em.desc()).all()
    
    # Calcular totais
    total_vales = len(vales)
    total_pallets_vales = sum(vale.quantidade_pallets for vale in vales)
    total_devolucoes = len(devolucoes)
    total_pallets_devolucoes = sum(dev.quantidade_pallets for dev in devolucoes)
    
    # Buscar dados para os filtros
    tipos_empresa = TipoEmpresa.query.all()
    
    # Buscar empresas (todas as empresas que aparecem em vales ou devoluções)
    empresas_ids = set()
    for vale in ValePallet.query.all():
        if vale.cliente_id:
            empresas_ids.add(vale.cliente_id)
        if vale.transportadora_id:
            empresas_ids.add(vale.transportadora_id)
        if vale.destinatario_id:
            empresas_ids.add(vale.destinatario_id)
    
    for dev in DevolucaoPallet.query.all():
        if dev.cliente_id:
            empresas_ids.add(dev.cliente_id)
        if dev.transportadora_id:
            empresas_ids.add(dev.transportadora_id)
        if dev.destinatario_id:
            empresas_ids.add(dev.destinatario_id)
    
    empresas = Empresa.query.filter(Empresa.id.in_(empresas_ids)).order_by(Empresa.razao_social).all() if empresas_ids else []
    
    # Buscar motoristas que têm vales ou devoluções
    motoristas_ids = set()
    for vale in ValePallet.query.all():
        if vale.motorista_id:
            motoristas_ids.add(vale.motorista_id)
    for dev in DevolucaoPallet.query.all():
        if dev.motorista_id:
            motoristas_ids.add(dev.motorista_id)
    
    motoristas = Motorista.query.filter(Motorista.id.in_(motoristas_ids)).order_by(Motorista.nome).all() if motoristas_ids else []
    
    return render_template('relatorios/movimentacao.html',
                         filtro_tipo_relatorio=filtro_tipo_relatorio,
                         vales=vales,
                         devolucoes=devolucoes,
                         total_vales=total_vales,
                         total_pallets_vales=total_pallets_vales,
                         total_devolucoes=total_devolucoes,
                         total_pallets_devolucoes=total_pallets_devolucoes,
                         tipos_empresa=tipos_empresa,
                         empresas=empresas,
                         motoristas=motoristas,
                         filtro_tipo=filtro_tipo,
                         filtro_empresa=filtro_empresa,
                         filtro_motorista=filtro_motorista,
                         filtro_documento=filtro_documento,
                         filtro_status=filtro_status)


@bp.route('/movimentacao/exportar-excel')
@login_required
def exportar_movimentacao_excel():
    """Exporta relatório de movimentação para Excel (respeitando filtros)"""
    from datetime import datetime
    from io import BytesIO
    from flask import send_file
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    # Obter filtros (mesmos do relatório)
    filtro_tipo_relatorio = request.args.get('tipo_relatorio', 'vales')
    filtro_tipo = request.args.get('tipo_empresa', '')
    filtro_empresa = request.args.get('empresa', '')
    filtro_motorista = request.args.get('motorista', '')
    filtro_documento = request.args.get('documento', '')
    filtro_status = request.args.get('status', '')
    
    vales = []
    devolucoes = []
    
    # Buscar Vales Pallet (mesma lógica do relatório)
    if filtro_tipo_relatorio in ['vales', 'consolidado']:
        query_vales = ValePallet.query
        
        if filtro_tipo:
            if filtro_tipo == 'cliente':
                query_vales = query_vales.join(ValePallet.cliente).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome == 'Cliente'
                )
            elif filtro_tipo == 'transportadora':
                query_vales = query_vales.join(ValePallet.transportadora).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome == 'Transportadora'
                )
            elif filtro_tipo == 'destinatario':
                query_vales = query_vales.join(ValePallet.destinatario).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome.in_(['Destinatário', 'Destinatario'])
                )
        
        if filtro_empresa:
            empresas_filtradas = Empresa.query.filter(
                or_(
                    Empresa.razao_social.ilike(f'%{filtro_empresa}%'),
                    Empresa.nome_fantasia.ilike(f'%{filtro_empresa}%')
                )
            ).all()
            empresas_ids = [e.id for e in empresas_filtradas]
            
            if empresas_ids:
                query_vales = query_vales.filter(
                    or_(
                        ValePallet.cliente_id.in_(empresas_ids),
                        ValePallet.transportadora_id.in_(empresas_ids),
                        ValePallet.destinatario_id.in_(empresas_ids)
                    )
                )
        
        if filtro_motorista:
            query_vales = query_vales.join(ValePallet.motorista).filter(
                Motorista.nome.ilike(f'%{filtro_motorista}%')
            )
        
        if filtro_documento:
            query_vales = query_vales.filter(
                ValePallet.numero_documento.ilike(f'%{filtro_documento}%')
            )
        
        if filtro_status:
            query_vales = query_vales.filter(ValePallet.status == filtro_status)
        
        vales = query_vales.order_by(ValePallet.data_criacao.desc()).all()
    
    # Buscar Devoluções (mesma lógica do relatório)
    if filtro_tipo_relatorio in ['devolucoes', 'consolidado']:
        query_devolucoes = DevolucaoPallet.query
        
        if filtro_tipo:
            if filtro_tipo == 'cliente':
                query_devolucoes = query_devolucoes.join(DevolucaoPallet.cliente).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome == 'Cliente'
                )
            elif filtro_tipo == 'transportadora':
                query_devolucoes = query_devolucoes.join(DevolucaoPallet.transportadora).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome == 'Transportadora'
                )
            elif filtro_tipo == 'destinatario':
                query_devolucoes = query_devolucoes.join(DevolucaoPallet.destinatario).join(Empresa.tipo_empresa).filter(
                    TipoEmpresa.nome.in_(['Destinatário', 'Destinatario'])
                )
        
        if filtro_empresa:
            empresas_filtradas = Empresa.query.filter(
                or_(
                    Empresa.razao_social.ilike(f'%{filtro_empresa}%'),
                    Empresa.nome_fantasia.ilike(f'%{filtro_empresa}%')
                )
            ).all()
            empresas_ids = [e.id for e in empresas_filtradas]
            
            if empresas_ids:
                query_devolucoes = query_devolucoes.filter(
                    or_(
                        DevolucaoPallet.cliente_id.in_(empresas_ids),
                        DevolucaoPallet.transportadora_id.in_(empresas_ids),
                        DevolucaoPallet.destinatario_id.in_(empresas_ids)
                    )
                )
        
        if filtro_motorista:
            query_devolucoes = query_devolucoes.join(DevolucaoPallet.motorista).filter(
                Motorista.nome.ilike(f'%{filtro_motorista}%')
            )
        
        if filtro_status:
            query_devolucoes = query_devolucoes.filter(DevolucaoPallet.status == filtro_status)
        
        devolucoes = query_devolucoes.order_by(DevolucaoPallet.criado_em.desc()).all()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # ABA 1: Vales Pallet (se houver)
    if vales and filtro_tipo_relatorio in ['vales', 'consolidado']:
        ws_vales = wb.active
        ws_vales.title = "Vales Pallet"
        
        # Cabeçalhos
        headers_vales = [
            "PIN", "Documento", "Cliente", "Destinatário", "Transportadora",
            "Motorista", "Placa", "Quantidade", "Devolvida", "Saldo",
            "Vencimento", "Data Criação", "Status"
        ]
        
        for col_num, header in enumerate(headers_vales, 1):
            cell = ws_vales.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Dados
        for row_num, vale in enumerate(vales, 2):
            ws_vales.cell(row=row_num, column=1, value=vale.pin)
            ws_vales.cell(row=row_num, column=2, value=vale.numero_documento)
            ws_vales.cell(row=row_num, column=3, value=vale.cliente.razao_social if vale.cliente else "")
            ws_vales.cell(row=row_num, column=4, value=vale.destinatario.razao_social if vale.destinatario else "")
            ws_vales.cell(row=row_num, column=5, value=vale.transportadora.razao_social if vale.transportadora else "")
            ws_vales.cell(row=row_num, column=6, value=vale.motorista.nome if vale.motorista else "")
            ws_vales.cell(row=row_num, column=7, value=vale.motorista.placa_caminhao if vale.motorista else "")
            ws_vales.cell(row=row_num, column=8, value=vale.quantidade_pallets)
            ws_vales.cell(row=row_num, column=9, value=vale.quantidade_devolvida)
            ws_vales.cell(row=row_num, column=10, value=vale.quantidade_pallets - vale.quantidade_devolvida)
            ws_vales.cell(row=row_num, column=11, value=vale.data_vencimento.strftime('%d/%m/%Y') if vale.data_vencimento else "")
            ws_vales.cell(row=row_num, column=12, value=vale.data_criacao.strftime('%d/%m/%Y %H:%M') if vale.data_criacao else "")
            ws_vales.cell(row=row_num, column=13, value=vale.get_status_display())
        
        # Ajustar largura
        for col in ws_vales.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_vales.column_dimensions[column].width = adjusted_width
    
    # ABA 2: Devoluções (se houver)
    if devolucoes and filtro_tipo_relatorio in ['devolucoes', 'consolidado']:
        if filtro_tipo_relatorio == 'consolidado' and vales:
            ws_dev = wb.create_sheet("Devoluções")
        else:
            ws_dev = wb.active
            ws_dev.title = "Devoluções"
        
        # Cabeçalhos
        headers_dev = [
            "PIN", "Cliente", "Destinatário", "Transportadora", "Motorista",
            "Placa", "Quantidade", "Data Agendamento", "Data Coleta",
            "Status"
        ]
        
        for col_num, header in enumerate(headers_dev, 1):
            cell = ws_dev.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Dados
        for row_num, dev in enumerate(devolucoes, 2):
            ws_dev.cell(row=row_num, column=1, value=dev.pin_devolucao)
            ws_dev.cell(row=row_num, column=2, value=dev.cliente.razao_social if dev.cliente else "")
            ws_dev.cell(row=row_num, column=3, value=dev.destinatario.razao_social if dev.destinatario else "")
            ws_dev.cell(row=row_num, column=4, value=dev.transportadora.razao_social if dev.transportadora else "")
            ws_dev.cell(row=row_num, column=5, value=dev.motorista.nome if dev.motorista else "")
            ws_dev.cell(row=row_num, column=6, value=dev.motorista.placa_caminhao if dev.motorista else "")
            ws_dev.cell(row=row_num, column=7, value=dev.quantidade_pallets)
            ws_dev.cell(row=row_num, column=8, value=dev.data_agendamento.strftime('%d/%m/%Y') if dev.data_agendamento else "")
            ws_dev.cell(row=row_num, column=9, value=dev.data_coleta.strftime('%d/%m/%Y %H:%M') if dev.data_coleta else "")
            ws_dev.cell(row=row_num, column=10, value=dev.get_status_display())
        
        # Ajustar largura
        for col in ws_dev.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_dev.column_dimensions[column].width = adjusted_width
    
    # Salvar em BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome do arquivo
    tipo_nome = {
        'vales': 'vales_pallet',
        'devolucoes': 'devolucoes',
        'consolidado': 'movimentacao_consolidada'
    }
    nome_arquivo = f"{tipo_nome.get(filtro_tipo_relatorio, 'relatorio')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=nome_arquivo
    )
